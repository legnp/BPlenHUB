import openpyxl
import os

portfolio_path = r"D:\BPlen HUB\v3\portfolio\portfolio_bplen.xlsx"

# Load the workbook (without read_only or data_only to preserve other sheets and formulas)
wb = openpyxl.load_workbook(portfolio_path)

if "Checkpoints" in wb.sheetnames:
    print("Checkpoints sheet already exists. Removing it to rewrite clean.")
    wb.remove(wb["Checkpoints"])

# Create the sheet
ws = wb.create_sheet("Checkpoints")

# Headers
headers = ["ServiceCode", "CheckpointId", "Order", "Title", "Type", "ReferenceId", "Description"]
ws.append(headers)

# Data rows
checkpoints_data = [
    # BPL-000 Onboarding
    ["BPL-000", "introducao", 1, "Introdução", "content", "welcome_video_01", "Conheça a visão da BPlen"],
    ["BPL-000", "check_in_survey", 2, "Check-in", "survey", "check_in", "Demandas e estado atual da jornada"],
    ["BPL-000", "sessao_onboarding", 3, "Sessão de Onboarding", "meeting", "onboarding", "Agende sua sessão individual de boas-vindas com nossos orientadores."],
    
    # BPL-001 Posicionamento Profissional
    ["BPL-001", "perfil-profissional", 1, "Perfil Profissional", "form", "perfil_profissional", "Preenchimento do Perfil Profissional para estruturação do posicionamento."],
    ["BPL-001", "agendar-orientacao", 2, "Agendar Orientação Individual", "meeting", "orientacao-individual-posicionamento", "Sessão individual com orientador especialista de 1h."],
    
    # BPL-002 Análise Comportamental (DISC)
    ["BPL-002", "realizar-disc", 1, "Responder Questionário DISC", "survey", "disc", "Formulário psicométrico oficial."],
    ["BPL-002", "agendar-devolutiva", 2, "Agendar Devolutiva Individual", "meeting", "devolutiva-analise-comportamental", "Sessão com orientador especialista de 1h."],
    
    # BPL-003 Plano de Carreira
    ["BPL-003", "plano-carreira-form", 1, "Plano de Carreira", "form", "plano_carreira", "Preenchimento do formulário do Plano de Carreira."],
    ["BPL-003", "agendar-devolutiva-plano", 2, "Agendar Devolutiva do Plano", "meeting", "devolutiva-plano-carreira", "Sessão individual de alinhamento e entrega do Plano de Carreira."],
    
    # BPL-004 Gestão e Desenvolvimento (GDC)
    ["BPL-004", "gdc-form", 1, "Gestão de Carreira", "form", "gdc", "Formulário de acompanhamento de Gestão e Desenvolvimento de Carreira."],
    ["BPL-004", "agendar-mentorias", 2, "Agendar Mentorias", "meeting", "gdc-mentorias", "Agendamento de sessões periódicas de mentoria."],
    
    # BPL-005 MentoCoach
    ["BPL-005", "mentocoach-form", 1, "Acompanhamento MentoCoach", "form", "mentocoach", "Formulário de autoavaliação e acompanhamento MentoCoach."],
    ["BPL-005", "agendar-sessoes", 2, "Agendar Sessões de Coaching", "meeting", "mentocoach-sessoes", "Agendamento das sessões individuais de coaching de alta performance."],
    
    # BPL-006 Offboarding
    ["BPL-006", "offboarding-survey", 1, "Avaliação Final", "survey", "offboarding_survey", "Pesquisa de encerramento e consolidação de resultados."]
]

for row in checkpoints_data:
    ws.append(row)

wb.save(portfolio_path)
print("Checkpoints sheet created and populated successfully in portfolio_bplen.xlsx!")
